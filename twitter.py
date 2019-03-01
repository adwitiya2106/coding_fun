from bs4 import BeautifulSoup as soup
from urllib.request import urlopen as ureq
from googlesearch import search
import csv
from datetime import datetime
#query = "twitter michigan ross mba student reviews"
import openpyxl
wb=openpyxl.load_workbook("C:\\Users\\adity\\OneDrive\\Desktop\\Target_Python\\SHarman\\new_list_college.xlsx")
ws=wb.worksheets[1]
col_count=0
for lc in range(2,4):
 try:   
    my_url='' 
    col_name=ws.cell(lc,1).value
    fcol_name="twitter "+col_name  
    print(fcol_name)
    for j in search(fcol_name, tld="com", num=1, stop=5):
        if ("?lang=en" in j):
            my_url=j
            break    
    if my_url:
        pass
    else:
        for j in search(fcol_name, tld="com", num=1, stop=5):
            if ("twitter" in j):
                my_url=j
                break
            
   
    print("Link")    
    my_page=ureq(my_url)
    page_read=my_page.read()
    my_page.close()
    
        

    html_read=soup(page_read,"html.parser")
    row_container=html_read.findAll("ul",{"class":"ProfileNav-list"})
    tweet_values=[]
    #print("College Name= "+col_name)
    print(my_url)
    for obj in row_container[0].find_all("li"):
      try:
        print(col_name)
        print(obj.a["title"])
        tweet_values.append(obj.a["title"].strip())
      except:
        pass
    #print("****************************************")
    '''
    tt=tweet_values[0].split()
    tcount=int(tt[0].replace(",",""))
    ttext=tt[1]
    ws.cell(lc,12).value=tcount
    ws.cell(lc,13).value=ttext
    
    tt=tweet_values[1].split()
    tcount=int(tt[0].replace(",",""))
    ttext=tt[1]
    ws.cell(lc,14).value=tcount
    ws.cell(lc,15).value=ttext
    
    tt=tweet_values[2].split()
    tcount=int(tt[0].replace(",",""))
    ttext=tt[1]
    ws.cell(lc,16).value=tcount
    ws.cell(lc,17).value=ttext
    
    tt=tweet_values[3].split()
    tcount=int(tt[0].replace(",",""))
    ttext=tt[1]
    ws.cell(lc,18).value=tcount
    ws.cell(lc,19).value=ttext
    '''
 except:
    print("In exception: ") 
    continue
#wb.save("C:\\Users\\adity\\OneDrive\\Desktop\\Target_Python\\SHarman\\new_list_college.xlsx")    
    