from bs4 import BeautifulSoup as soup
from urllib.request import urlopen as ureq
from googlesearch import search
import csv
from selenium import webdriver
from datetime import datetime
#query = "twitter michigan ross mba student reviews"
import openpyxl
wb=openpyxl.load_workbook("C:\\Users\\adity\\OneDrive\\Desktop\\Target_Python\\SHarman\\new_list_college.xlsx")
ws=wb.worksheets[1]

driver=webdriver.Chrome()
driver.get('https://twitter.com/login')
username_box=driver.find_element_by_css_selector("#page-container > div > div.signin-wrapper > form > fieldset > div:nth-child(2) > input")
password_box=driver.find_element_by_css_selector("#page-container > div > div.signin-wrapper > form > fieldset > div:nth-child(3) > input")
username_box.send_keys("addystar@gmail.com")
password_box.send_keys("sum#1234")

login_button=driver.find_element_by_css_selector("#page-container > div > div.signin-wrapper > form > div.clearfix > button")
login_button.click()

search_box=driver.find_element_by_css_selector('#search-query')
#search_box.send_keys("Test")
#search_box.submit()
#time.sleep(1.5)
#latest_link=driver.find_element_by_css_selector('#page-container > div.SearchNavigation > div.AdaptiveFiltersBar > div > ul > li:nth-child(2) > a')    
for lc in range(2,505):
 try:
    #print("Inside FOR loop")
    posts=[]
    col_name=ws.cell(lc,1).value
    search_box.clear()
    search_box.send_keys(col_name) 
    search_box.submit()
    time.sleep(3)
    #latest_link=driver.find_element_by_css_selector('#page-container > div.SearchNavigation > div.AdaptiveFiltersBar > div > ul > li:nth-child(2) > a')
    #latest_link.click()
    #time.sleep(3)
    SCROLL_PAUSE_TIME=2
    for i in range(20):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            # Wait to load page
        time.sleep(SCROLL_PAUSE_TIME)     
    
    #posts=driver.find_elements_by_class_name("stream-item-header")
    time.sleep(SCROLL_PAUSE_TIME+4)
    posts=driver.find_elements_by_class_name("FullNameGroup")
    print(lc,col_name)
    print("Total Tweets: ",len(posts))
    print("***************************")
    ws.cell(lc,12).value=len(posts)
    if ((lc==100) or (lc==200) or (lc==300) or (lc==400)):
        wb.save("C:\\Users\\adity\\OneDrive\\Desktop\\Target_Python\\SHarman\\new_list_college.xlsx")
 except:
    continue  
wb.save("C:\\Users\\adity\\OneDrive\\Desktop\\Target_Python\\SHarman\\new_list_college.xlsx")
   