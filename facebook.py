from googlesearch import search
import openpyxl
work_book=openpyxl.load_workbook("C:\\Users\\adity\\OneDrive\\Desktop\\Target_Python\\SHarman\\ACCCSB_FINAL_LIST.xlsx")
ws=work_book.worksheets[1]
col_name=[]
for lc in range(1,ws.max_row+1):
    #col_name.append(ws.cell(lc,1).value)
    my_url=[]
    query='Facebook page '+ ws.cell(lc,1).value
    for j in search(query, tld="com", num=1, stop=5, pause=2): 
        if ('facebook' in j):
            my_url.append(j)
    ws.cell(lc,3).value=min(my_url,key=len)  
    print(ws.cell(lc,1).value+' '+ws.cell(lc,3).value)         
work_book.save("C:\\Users\\adity\\OneDrive\\Desktop\\Target_Python\\SHarman\\ACCCSB_FINAL_LIST.xlsx")