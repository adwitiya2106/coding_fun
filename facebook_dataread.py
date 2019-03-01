import openpyxl
from selenium import webdriver
import time
from statistics import mean


start_time=time.time()
work_book=openpyxl.load_workbook("C:\\Users\\adity\\OneDrive\\Desktop\\Target_Python\\SHarman\\new_list_college.xlsx")
ws=work_book.worksheets[1]

#login
chrome_options = webdriver.ChromeOptions()
prefs = {"profile.default_content_setting_values.notifications" : 2}
chrome_options.add_experimental_option("prefs",prefs)
driver = webdriver.Chrome(chrome_options=chrome_options)
driver.get("https://www.facebook.com")
#login
login=driver.find_element_by_css_selector('#email')
pass_word=driver.find_element_by_css_selector('#pass')
login_button=driver.find_element_by_css_selector('#u_0_2')
login.send_keys('adwitiya.kr@gmail.com')
pass_word.send_keys('Sum#1234')                                              
login_button.click()
    

for lc in range(504,505):  
    try:
        num_likes=''
        num_follow=''
        min_like=''
        max_like=''
        mean_like=''
        rating=''
        total_people=''
        fb_url=ws.cell(lc,2).value


        driver.get(fb_url)
        #Public Home page reached
        page_like_follow=driver.find_elements_by_class_name('_4bl9') #likes and follow
        page_like_follow_count=[]
        for i in (page_like_follow):
             page_like_follow_count.append(i.text)
        for i in page_like_follow_count:
            if 'people like' in i:
                num_likes=i.replace('people like this','').strip()
            elif 'people follow' in i:    
                num_follow=i.replace('people follow this','').strip()

        #name=driver.find_elements_by_class_name('_4arz')

        SCROLL_PAUSE_TIME=0.5
        for i in range(20):
            # Scroll down to bottom
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            # Wait to load page
            time.sleep(SCROLL_PAUSE_TIME)

        name=driver.find_elements_by_class_name('_4arz')
        name_text=[]
        for i in (name):
            name_text.append(i.text)
        name_text=list(map(int,name_text))    
        min_like=min(name_text)
        max_like=max(name_text)
        mean_like=mean(name_text)

        name=driver.find_elements_by_class_name('_2w0a')
        rating=name[0].text.replace("out of 5","").strip()
        name=driver.find_elements_by_class_name('_2w0b')
        popu=name[0].text
        popu=popu.split(sep=' ')
        total_people=popu[-2]
        
        ws.cell(lc,3).value=int(num_likes.replace(",",""))
        ws.cell(lc,4).value=int(num_follow.replace(",",""))
        ws.cell(lc,5).value=min_like
        ws.cell(lc,6).value=max_like
        ws.cell(lc,7).value=mean_like
        ws.cell(lc,8).value=float(rating)
        ws.cell(lc,9).value=int(total_people)
        
        print("FINISHED: "+fb_url)
        
    except:
        ws.cell(lc,10).value='Exception'
        ws.cell(lc,3).value=int(num_likes.replace(",","") or 0)
        ws.cell(lc,4).value=int(num_follow.replace(",","") or 0)
        ws.cell(lc,5).value=min_like
        ws.cell(lc,6).value=max_like
        ws.cell(lc,7).value=mean_like
        ws.cell(lc,8).value=float(rating or 0)
        ws.cell(lc,9).value=int(total_people.replace(",","") or 0)
        print("FINISHED: "+fb_url)
        continue
work_book.save("C:\\Users\\adity\\OneDrive\\Desktop\\Target_Python\\SHarman\\new_list_college.xlsx")
print("Total Time:",time.time()-start_time)    