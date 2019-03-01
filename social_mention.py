import openpyxl
from selenium import webdriver
from bs4 import BeautifulSoup as soup
from urllib.request import urlopen as ureq
import time

work_book=openpyxl.load_workbook("C:\\Users\\adity\\OneDrive\\Desktop\\Target_Python\\SHarman\\new_list_college.xlsx")
ws=work_book.worksheets[0]
col_name=[]  
wb=webdriver.Chrome()
wb.get("http://socialmention.com/")
sbox=wb.find_element_by_css_selector('#search > table > tbody > tr > td:nth-child(1) > input[type="text"]')
#col_list=["michigan ross mba","harvard business school mba","ubuffalo mba"]
for lc in range(2,6):
    try:  
     sbox.send_keys(ws.cell(lc,1).value)
     sbox.submit() 
    
     spos=wb.find_element_by_css_selector('#top_keywords > table > tbody > tr:nth-child(1) > td:nth-child(3)')
     sneg=wb.find_element_by_css_selector('#top_keywords > table > tbody > tr:nth-child(3) > td:nth-child(3)')
     sneutral=wb.find_element_by_css_selector('#top_keywords > table > tbody > tr:nth-child(2) > td:nth-child(3)')
     strength=wb.find_element_by_css_selector('#score_strength > div.score')
     passion=wb.find_element_by_css_selector('#score_passion > div.score')
     reach=wb.find_element_by_css_selector('#score_reach > div.score')
    
     print(ws.cell(lc,1).value)
     print("Positive: ",int(spos.text))
     print("Negative: ",int(sneg.text))
     print("Neutral: ",int(sneutral.text))
     print("strength: ",int(strength.text.replace("%","")))
     print("Passion: ",int(passion.text.replace("%","")))
     print("Reach: ",int(reach.text.replace("%","")))
     print("***********************************************")
     #ws.cell(lc,2).value=int(spos.text)
     #ws.cell(lc,3).value=int(sneg.text)
     #ws.cell(lc,4).value=int(sneutral.text)
     #ws.cell(lc,5).value=int(strength.text.replace("%",""))
     #ws.cell(lc,6).value=int(passion.text.replace("%",""))
     #ws.cell(lc,7).value=int(reach.text.replace("%",""))
    
    except:
       pass 
    sbox=wb.find_element_by_css_selector('#search_top > form > input.search_box')
    sbox.clear()
#work_book.save("C:\\Users\\adity\\OneDrive\\Desktop\\Target_Python\\SHarman\\new_list_college.xlsx")   